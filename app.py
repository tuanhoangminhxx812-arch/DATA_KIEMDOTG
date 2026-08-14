import os
import sys

# Tự động phát hiện khi chạy trên Streamlit Cloud hoặc qua lệnh streamlit run
IS_STREAMLIT = any('streamlit' in str(arg).lower() for arg in sys.argv) or 'STREAMLIT_SERVER_PORT' in os.environ or 'STREAMLIT_RUN' in os.environ

if IS_STREAMLIT:
    # Nếu chạy qua Streamlit Cloud, tự động chuyển sang giao diện Streamlit
    import app_streamlit
else:
    import shutil
    import openpyxl
    import re
    from flask import Flask, request, jsonify, render_template, send_file
    import loc_lech_taikhoan

    app = Flask(__name__)
    app.config['JSON_AS_ASCII'] = False

    # Thư mục chứa các file upload và kết quả
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    # Trạng thái hiện tại của ứng dụng
    app_state = {
        "original_filename": None,
        "original_filepath": None,
        "analyzed_filepath": None,
        "analysis_data": None
    }


def get_safe_filename(orig_name):
    if not orig_name:
        return "data_upload.xlsx"
    name = os.path.basename(orig_name)
    name = re.sub(r'[\/\\:\*\?"<>\|]', '_', name).strip()
    if not name or name == '.xlsx':
        name = "data_upload.xlsx"
    return name

def load_and_calculate_data(file_path):
    """
    Đọc dữ liệu từ file Excel đã được đối chiếu để trả về cấu trúc JSON cho Frontend.
    Sử dụng openpyxl với read_only=True để tối ưu bộ nhớ và tốc độ xử lý.
    """
    # Tìm backup path để đọc dữ liệu gốc
    backup_path = None
    if app_state["original_filepath"]:
        orig_path = app_state["original_filepath"]
        orig_dir = os.path.dirname(orig_path)
        orig_name = os.path.basename(orig_path)
        orig_without_ext, orig_ext = os.path.splitext(orig_name)
        candidate = os.path.join(orig_dir, f"{orig_without_ext}_backup{orig_ext}")
        if os.path.exists(candidate):
            backup_path = candidate
            
    source_path = backup_path if backup_path else file_path
    print(f"Reading source data for Table A from: {source_path}")
    wb_source = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    source_sheet_name, header_row, indices = loc_lech_taikhoan.detect_sheet_and_headers(wb_source)
    if not source_sheet_name:
        wb_source.close()
        wb_source = openpyxl.load_workbook(source_path, data_only=False, read_only=True)
        source_sheet_name, header_row, indices = loc_lech_taikhoan.detect_sheet_and_headers(wb_source)

    if not source_sheet_name:
        wb_source.close()
        raise ValueError("Không tìm thấy sheet dữ liệu gốc hợp lệ hoặc lỗi cấu trúc cột trong file Excel.")
        
    sheet = wb_source[source_sheet_name]
    
    acc_idx = indices['account']
    deb_idx = indices['debit']
    cred_idx = indices['credit']
    creator_idx = indices['creator']
        
    intermediate_accounts = [
        '11310000000', '14190000000', '15110000000', '15190000000', '24190000000',
        '33191000000', '33192000000', '33193000000', '33194000000', '33195000000',
        '33196100000', '33198000000'
    ]
    
    # Khởi tạo Bảng A
    bang_a_data = {acc: {"account": acc, "debit": 0.0, "credit": 0.0, "diff": 0.0, "status": "Khớp"} for acc in intermediate_accounts}
    all_creators = set()
    
    # Duyệt qua các dòng dữ liệu gốc từ dòng ngay sau dòng tiêu đề
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) <= acc_idx or row[acc_idx] is None:
            continue
        acc = str(row[acc_idx]).strip().strip('="\'')
        if not acc:
            continue
        
        # Thu thập Người tạo (Creator) xuất hiện trong hệ thống
        if len(row) > creator_idx and row[creator_idx] is not None:
            c_val = str(row[creator_idx]).strip().strip('="\'')
            if c_val:
                all_creators.add(c_val)
            
        if acc in bang_a_data:
            deb_val = loc_lech_taikhoan.parse_float(row[deb_idx]) if len(row) > deb_idx else 0.0
            cred_val = loc_lech_taikhoan.parse_float(row[cred_idx]) if len(row) > cred_idx else 0.0
            
            bang_a_data[acc]["debit"] += deb_val
            bang_a_data[acc]["credit"] += cred_val

    wb_source.close()

    # Tính toán chênh lệch và trạng thái Bảng A
    for acc in intermediate_accounts:
        item = bang_a_data[acc]
        item["diff"] = item["debit"] - item["credit"]
        item["status"] = "Lệch số" if abs(round(item["diff"], 2)) != 0 else "Khớp"

    # 2. Đọc sheet Data_Unpaired để lấy các giao dịch chưa khớp và tính toán Bảng B
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if 'Data_Unpaired' not in wb.sheetnames:
        wb.close()
        raise ValueError("Ứng dụng chưa được Phân Tích Đối Chiếu. Vui lòng bấm nút Phân Tích.")
        
    ws_data = wb['Data_Unpaired']
    unpaired_transactions = []
    
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
            
        date_val = row[5] if len(row) > 5 else ''
        if date_val and hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val or '')
            
        if len(row) > 1 and row[1]:
            all_creators.add(str(row[1]).strip())
            
        row_num = 0
        if len(row) > 7 and row[7] is not None:
            try:
                row_num = int(row[7])
            except (ValueError, TypeError):
                row_num = 0

        deb_num = 0.0
        if len(row) > 2 and row[2] is not None:
            try:
                deb_num = float(row[2])
            except (ValueError, TypeError):
                deb_num = 0.0

        cred_num = 0.0
        if len(row) > 3 and row[3] is not None:
            try:
                cred_num = float(row[3])
            except (ValueError, TypeError):
                cred_num = 0.0

        unpaired_transactions.append({
            "account": str(row[0]).strip(),
            "creator": str(row[1] or 'UNKNOWN').strip(),
            "deb": deb_num,
            "cred": cred_num,
            "trans_num": str(row[4] or '') if len(row) > 4 else '',
            "date": date_str,
            "desc": str(row[6] or '') if len(row) > 6 else '',
            "row_num": row_num,
            "updater": str(row[8] or '') if len(row) > 8 else ''
        })
        
    wb.close()
        
    # Nhóm theo Tài khoản và Người tạo cho Bảng B
    bang_b_data = {}
    for acc in intermediate_accounts:
        bang_b_data[acc] = {}
        for creator in all_creators:
            bang_b_data[acc][creator] = {"creator": creator, "debit": 0.0, "credit": 0.0, "diff": 0.0, "status": ""}
        
    for tx in unpaired_transactions:
        acc = tx["account"]
        creator = tx["creator"]
        if acc not in bang_b_data:
            bang_b_data[acc] = {}
        if creator not in bang_b_data[acc]:
            bang_b_data[acc][creator] = {"creator": creator, "debit": 0.0, "credit": 0.0, "diff": 0.0, "status": ""}
            
        bang_b_data[acc][creator]["debit"] += tx["deb"]
        bang_b_data[acc][creator]["credit"] += tx["cred"]

    # Tính chênh lệch và trạng thái Bảng B
    for acc in intermediate_accounts:
        for creator, item in bang_b_data[acc].items():
            item["diff"] = item["debit"] - item["credit"]
            if abs(round(item["diff"], 2)) != 0:
                item["status"] = "Lệch số"
            elif item["debit"] > 0 or item["credit"] > 0:
                item["status"] = "Khớp"
            else:
                item["status"] = ""
            
    bang_b_list = {
        acc: sorted(list(creators.values()), key=lambda x: x["creator"]) 
        for acc, creators in bang_b_data.items()
    }
    
    # 3. Phân nhóm các giao dịch chưa khớp cho Bảng Chi Tiết
    chi_tiet_data = {}
    for acc in intermediate_accounts:
        chi_tiet_data[acc] = []
        
    for tx in unpaired_transactions:
        acc = tx["account"]
        if acc in chi_tiet_data:
            chi_tiet_data[acc].append(tx)
            
    for acc in intermediate_accounts:
        chi_tiet_data[acc].sort(key=lambda x: x["row_num"])
        
    # 4. Tính toán các Thẻ Thống kê KPI
    total_unbalanced_accs = sum(1 for acc in intermediate_accounts if bang_a_data[acc]["status"] == "Lệch số")
    total_unpaired_debit = sum(tx["deb"] for tx in unpaired_transactions)
    total_unpaired_credit = sum(tx["cred"] for tx in unpaired_transactions)
    total_discrepancy = total_unpaired_debit - total_unpaired_credit
    
    kpi_cards = {
        "total_unbalanced_accs": total_unbalanced_accs,
        "total_unpaired_debit": total_unpaired_debit,
        "total_unpaired_credit": total_unpaired_credit,
        "total_discrepancy": total_discrepancy
    }
    
    return {
        "kpi_cards": kpi_cards,
        "doi_chieu_bang_a": list(bang_a_data.values()),
        "doi_chieu_bang_b": bang_b_list,
        "chi_tiet_giao_dich": chi_tiet_data
    }

@app.route('/')
def index():
    """Trang chủ hiển thị ứng dụng"""
    return render_template('index.html')

@app.route('/api/state', methods=['GET'])
def get_state():
    """Lấy trạng thái hiện tại của ứng dụng"""
    return jsonify({
        "original_filename": app_state["original_filename"],
        "has_analyzed": app_state["analyzed_filepath"] is not None,
        "kpi_cards": app_state["analysis_data"]["kpi_cards"] if app_state["analysis_data"] else None
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """API tải lên file Excel mới"""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "Không tìm thấy file gửi lên"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "Tên file rỗng"}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({"success": False, "message": "Định dạng file không hợp lệ, vui lòng chọn file Excel .xlsx"}), 400
        
    try:
        import time
        timestamp = int(time.time() * 1000)
        safe_name = get_safe_filename(file.filename)
        name_without_ext, ext = os.path.splitext(safe_name)
        
        unique_filename = f"{name_without_ext}_{timestamp}{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(filepath)
        
        # Đọc thử file để kiểm tra cấu trúc sheet nguồn động
        wb = openpyxl.load_workbook(filepath, read_only=True)
        source_sheet_name, header_row, indices = loc_lech_taikhoan.detect_sheet_and_headers(wb)
        wb.close()
        
        if not source_sheet_name:
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({
                "success": False, 
                "message": "Cấu trúc file không đúng! File phải chứa một sheet sổ cái GL hợp lệ có các cột tối thiểu: 'Tài khoản', 'Nợ quy đổi', 'Có quy đổi', 'Nội dung', 'Người tạo'."
            }), 400
            
        # Tạo bản sao lưu backup ngay lập tức tại UPLOAD_FOLDER với tên duy nhất
        backup_path = os.path.join(UPLOAD_FOLDER, f"{name_without_ext}_{timestamp}_backup{ext}")
        shutil.copy2(filepath, backup_path)
        
        # Cập nhật trạng thái
        app_state["original_filename"] = file.filename
        app_state["original_filepath"] = filepath
        app_state["analyzed_filepath"] = None
        app_state["analysis_data"] = None
        
        return jsonify({
            "success": True, 
            "message": f"Tải file '{file.filename}' thành công! Sẵn sàng để Phân Tích Đối Chiếu."
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Lỗi xử lý file upload: {str(e)}"}), 500

@app.route('/api/analyze', methods=['POST'])
def analyze_file():
    """API chạy thuật toán đối chiếu dữ liệu"""
    if not app_state["original_filepath"]:
        return jsonify({"success": False, "message": "Vui lòng tải lên file Excel trước"}), 400
        
    try:
        # Gọi module thuật toán đối chiếu của loc_lech_taikhoan
        saved_path = loc_lech_taikhoan.main(app_state["original_filepath"])
        
        if not saved_path or not os.path.exists(saved_path):
            return jsonify({"success": False, "message": "Thuật toán chạy thành công nhưng không thể lưu file kết quả. Vui lòng kiểm tra quyền ghi hoặc đóng Excel đang mở."}), 500
            
        # Cập nhật đường dẫn file kết quả và load dữ liệu tính toán
        app_state["analyzed_filepath"] = saved_path
        app_state["analysis_data"] = load_and_calculate_data(saved_path)
        
        return jsonify({
            "success": True,
            "message": "Phân tích đối chiếu hoàn tất!",
            "data": app_state["analysis_data"]
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Lỗi khi chạy phân tích đối chiếu: {str(e)}"}), 500

@app.route('/api/data', methods=['GET'])
def get_data():
    """API lấy dữ liệu kết quả đối chiếu hiện tại"""
    if not app_state["analysis_data"]:
        return jsonify({"success": False, "message": "Chưa có dữ liệu phân tích"}), 400
    return jsonify({
        "success": True,
        "data": app_state["analysis_data"]
    })

@app.route('/api/export', methods=['GET'])
def export_file():
    """API xuất dữ liệu / tải file kết quả đối chiếu Excel về máy"""
    if not app_state["analyzed_filepath"] or not os.path.exists(app_state["analyzed_filepath"]):
        return jsonify({"success": False, "message": "Chưa có file kết quả phân tích để tải về"}), 400
        
    try:
        return send_file(
            app_state["analyzed_filepath"],
            as_attachment=True,
            download_name=f"KetQua_KiemDo_{app_state['original_filename']}"
        )
    except Exception as e:
        return jsonify({"success": False, "message": f"Lỗi tải file: {str(e)}"}), 500

@app.route('/api/export_filtered', methods=['GET'])
def export_filtered_file():
    """API xuất dữ liệu / tải file chi tiết các giao dịch lệch theo tài khoản và người tạo"""
    if not app_state["analyzed_filepath"] or not os.path.exists(app_state["analyzed_filepath"]):
        return jsonify({"success": False, "message": "Chưa có file kết quả phân tích để xuất"}), 400
        
    account = request.args.get('account', '').strip()
    creator = request.args.get('creator', '').strip()
    
    if not account:
        return jsonify({"success": False, "message": "Thiếu thông tin tài khoản cần xuất"}), 400
        
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Load và lọc dữ liệu từ file Data_Unpaired
        wb = openpyxl.load_workbook(app_state["analyzed_filepath"], data_only=True)
        if 'Data_Unpaired' not in wb.sheetnames:
            wb.close()
            return jsonify({"success": False, "message": "Không tìm thấy sheet dữ liệu đối chiếu"}), 400
            
        ws_data = wb['Data_Unpaired']
        rows = []
        
        # Đọc dữ liệu
        for row in ws_data.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            acc = str(row[0]).strip()
            cre = str(row[1] or 'UNKNOWN').strip()
            
            if acc == account:
                if not creator or cre == creator:
                    rows.append({
                        "row_num": int(row[7]),
                        "trans_num": str(row[4] or ''),
                        "date": str(row[5] or ''),
                        "deb": float(row[2] or 0),
                        "cred": float(row[3] or 0),
                        "creator": cre,
                        "updater": str(row[8] or '') if len(row) > 8 else '',
                        "desc": str(row[6] or '')
                    })
        wb.close()
        
        # Tạo workbook xuất mới
        wb_export = openpyxl.Workbook()
        ws = wb_export.active
        ws.title = "ChiTiet_Lech"
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_family = 'Segoe UI'
        title_font = Font(name=font_family, size=14, bold=True, color='1B365D')
        sub_font = Font(name=font_family, size=11, italic=True, color='595959')
        header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
        bold_font = Font(name=font_family, size=11, bold=True, color='000000')
        normal_font = Font(name=font_family, size=11, color='000000')
        red_font = Font(name=font_family, size=11, bold=True, color='C00000')
        
        header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid') # Navy Blue
        zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') # Zebra row
        red_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # Warning red
        
        thin_border_side = Side(border_style='thin', color='BFBFBF')
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        double_bottom_border = Border(
            left=thin_border_side, right=thin_border_side,
            top=thin_border_side,
            bottom=Side(border_style='double', color='1F497D')
        )
        
        # Tiêu đề báo cáo
        ws['A1'] = "BÁO CÁO CHI TIẾT CÁC GIAO DỊCH CHƯA ĐỐI ỨNG GÂY LỆCH"
        ws['A1'].font = title_font
        
        filter_text = f"Tài khoản trung gian: {account}"
        if creator:
            filter_text += f"   |   Người tạo (Creator): {creator}"
        ws['A2'] = filter_text
        ws['A2'].font = Font(name=font_family, size=11, bold=True, color='1B365D')
        
        ws['A3'] = "Trích xuất tự động từ Hệ thống Kiểm dò Tài khoản Trung Gian Điện lực Vũng Tàu"
        ws['A3'].font = sub_font
        
        # Tiêu đề bảng
        headers = ["Dòng gốc (Row)", "Số chứng từ", "Ngày giao dịch", "Số tiền Nợ", "Số tiền Có", "Người tạo (Creator)", "Người cập nhật", "Nội dung giao dịch"]
        header_row_num = 5
        ws.row_dimensions[header_row_num].height = 25
        
        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row_num, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Ghi các dòng dữ liệu
        current_row = 6
        for idx, r_data in enumerate(rows):
            ws.row_dimensions[current_row].height = 20
            
            c_row = ws.cell(row=current_row, column=1, value=r_data['row_num'])
            c_row.alignment = Alignment(horizontal='center')
            c_row.font = bold_font
            c_row.border = thin_border
            
            c_trans = ws.cell(row=current_row, column=2, value=r_data['trans_num'])
            c_trans.alignment = Alignment(horizontal='center')
            c_trans.font = normal_font
            c_trans.border = thin_border
            
            c_date = ws.cell(row=current_row, column=3, value=r_data['date'])
            c_date.alignment = Alignment(horizontal='center')
            c_date.font = normal_font
            c_date.border = thin_border
            
            c_deb = ws.cell(row=current_row, column=4, value=r_data['deb'])
            c_deb.number_format = '#,##0'
            c_deb.font = normal_font
            c_deb.border = thin_border
            
            c_cred = ws.cell(row=current_row, column=5, value=r_data['cred'])
            c_cred.number_format = '#,##0'
            c_cred.font = normal_font
            c_cred.border = thin_border
            
            c_cre = ws.cell(row=current_row, column=6, value=r_data['creator'])
            c_cre.font = bold_font
            c_cre.border = thin_border
            
            c_upd = ws.cell(row=current_row, column=7, value=r_data['updater'])
            c_upd.font = normal_font
            c_upd.border = thin_border
            
            c_desc = ws.cell(row=current_row, column=8, value=r_data['desc'])
            c_desc.font = normal_font
            c_desc.border = thin_border
            
            # Zebra striping
            if idx % 2 == 1:
                for c_idx in range(1, 9):
                    ws.cell(row=current_row, column=c_idx).fill = zebra_fill
                    
            current_row += 1
            
        # Dòng tổng cộng
        if len(rows) > 0:
            ws.row_dimensions[current_row].height = 22
            ws.cell(row=current_row, column=1, value="Tổng cộng chưa khớp").font = bold_font
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=current_row, column=1).border = double_bottom_border
            
            for c_idx in range(2, 4):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.border = double_bottom_border
                cell.fill = zebra_fill
                
            c_tot_deb = ws.cell(row=current_row, column=4, value=f"=SUM(D6:D{current_row-1})")
            c_tot_deb.number_format = '#,##0'
            c_tot_deb.font = bold_font
            c_tot_deb.border = double_bottom_border
            c_tot_deb.fill = zebra_fill
            
            c_tot_cred = ws.cell(row=current_row, column=5, value=f"=SUM(E6:E{current_row-1})")
            c_tot_cred.number_format = '#,##0'
            c_tot_cred.font = bold_font
            c_tot_cred.border = double_bottom_border
            c_tot_cred.fill = zebra_fill
            
            ws.cell(row=current_row, column=6, value="Chênh lệch (Nợ-Có):").font = bold_font
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=current_row, column=6).border = double_bottom_border
            ws.cell(row=current_row, column=6).fill = zebra_fill
            
            c_diff = ws.cell(row=current_row, column=7, value=f"=D{current_row}-E{current_row}")
            c_diff.number_format = '#,##0'
            c_diff.font = red_font
            c_diff.fill = red_fill
            c_diff.border = double_bottom_border
            
            ws.cell(row=current_row, column=8).border = double_bottom_border
            ws.cell(row=current_row, column=8).fill = zebra_fill
            
        # Tự động căn chỉnh độ rộng cột
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.row in [1, 2, 3] or (cell.value and str(cell.value).startswith('=')):
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.column_dimensions['H'].width = 50 # Diễn giải cho rộng ra
        
        # Lưu file xuất
        export_filename = f"ChiTiet_Lech_{account}"
        if creator:
            export_filename += f"_{creator}"
        export_filename += ".xlsx"
        
        export_path = os.path.join(UPLOAD_FOLDER, export_filename)
        wb_export.save(export_path)
        wb_export.close()
        
        return send_file(
            export_path,
            as_attachment=True,
            download_name=export_filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Lỗi xuất file chi tiết: {str(e)}"}), 500

if __name__ == '__main__':
    if not IS_STREAMLIT:
        app.run(host='0.0.0.0', port=5000, debug=True)

